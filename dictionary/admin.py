# dictionary/admin.py

from django.contrib import admin
from django import forms
from .models import DictionaryEntry, Tag, DictionaryTag, FavoriteEntry, UserProfile, CardSet, FavoriteCardSet, Article, VocabularyTopic, FavoriteVocabularyTopic
import openpyxl
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path
from django.shortcuts import render
from django.contrib import messages

# Форма для загрузки Excel-файла
class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(label="Выберите Excel-файл")

@admin.register(DictionaryEntry)
class DictionaryEntryAdmin(admin.ModelAdmin):
    list_display = ('eng_term', 'rus_term', 'transcription', 'description', 'created_at')
    search_fields = ('eng_term', 'rus_term')
    list_filter = ('created_at',)
    actions = ['export_to_excel']  # Оставляем только экспорт как действие

    # Добавляем кастомный URL для импорта
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='dictionary_entry_import_excel'),
        ]
        return custom_urls + urls

    # Представление для импорта
    def import_excel_view(self, request):
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    # Чтение Excel-файла
                    wb = openpyxl.load_workbook(excel_file)
                    sheet = wb.active

                    # Пропускаем первую строку (заголовки)
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        eng_term = row[0]  # Английское слово (столбец A)
                        rus_term = row[1]  # Русский перевод (столбец B)
                        transcription = row[2]  # Транскрипция (столбец C)
                        description = row[3]  # Описание на английском (столбец D)

                        # Пропускаем пустые строки
                        if not eng_term or not rus_term:
                            continue

                        # Создаём или обновляем запись
                        DictionaryEntry.objects.update_or_create(
                            eng_term=eng_term,
                            defaults={
                                'rus_term': rus_term,
                                'transcription': transcription if transcription else '',
                                'description': description if description else ''
                            }
                        )

                    self.message_user(request, "Данные успешно импортированы из Excel!", messages.SUCCESS)
                    return HttpResponseRedirect(request.get_full_path())
                except Exception as e:
                    self.message_user(request, f"Ошибка при импорте: {str(e)}", messages.ERROR)
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'title': 'Импортировать записи из Excel',
            'opts': self.model._meta,
            'has_permission': True,
        }
        return render(request, 'admin/import_excel_form.html', context)

    def export_to_excel(self, request, queryset):
        # Создаём новый Excel-файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dictionary Entries"

        # Добавляем заголовки
        headers = ["СЛОВО НА АНГЛ.", "ПЕРЕВОД", "Транскрипция", "Описание на англ."]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header

        # Заполняем данными
        for row_num, entry in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1).value = entry.eng_term
            ws.cell(row=row_num, column=2).value = entry.rus_term
            ws.cell(row=row_num, column=3).value = entry.transcription
            ws.cell(row=row_num, column=4).value = entry.description

        # Настраиваем ответ для скачивания файла
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="dictionary_entries_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        # Сохраняем файл в ответ
        wb.save(response)
        return response

    export_to_excel.short_description = "Экспортировать в Excel"

    # Добавляем кнопку "Импортировать из Excel" на страницу списка
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context=extra_context)

@admin.register(Tag)
class TagAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)

@admin.register(DictionaryTag)
class DictionaryTagAdmin(admin.ModelAdmin):
    list_display = ('entry', 'tag')
    list_filter = ('tag',)

@admin.register(FavoriteEntry)
class FavoriteEntryAdmin(admin.ModelAdmin):
    list_display = ('user', 'entry')
    list_filter = ('user',)

@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = ('user', 'group')
    list_filter = ('group',)

@admin.register(CardSet)
class CardSetAdmin(admin.ModelAdmin):
    list_display = ('name', 'user')
    list_filter = ('user',)

@admin.register(FavoriteCardSet)
class FavoriteCardSetAdmin(admin.ModelAdmin):
    list_display = ('user', 'card_set')
    list_filter = ('user',)

@admin.register(Article)
class ArticleAdmin(admin.ModelAdmin):
    list_display = ('title', 'created_at')
    list_filter = ('created_at',)

@admin.register(VocabularyTopic)
class VocabularyTopicAdmin(admin.ModelAdmin):
    list_display = ('name', 'user', 'is_default')
    list_filter = ('user', 'is_default')

@admin.register(FavoriteVocabularyTopic)
class FavoriteVocabularyTopicAdmin(admin.ModelAdmin):
    list_display = ('user', 'topic')
    list_filter = ('user',)